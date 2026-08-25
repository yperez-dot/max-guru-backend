#!/usr/bin/env python3
"""Phase 2: parse SoB text fields and diff against Max plan-data."""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

HTML = Path("/workspace/artifacts/max-demo-FINAL-v7.html")
EXTRACTED = Path("/workspace/artifacts/reports/sob_extracted.json")
OUT_CSV = Path("/workspace/artifacts/reports/sob-phase2-diff.csv")
OUT_XLSX = Path("/workspace/artifacts/reports/sob-phase2-corrections.xlsx")
OUT_JSON = Path("/workspace/artifacts/reports/sob-phase2-fields.json")


def load_plans() -> list[dict]:
    text = HTML.read_text(encoding="utf-8")
    m = re.search(
        r'<script id="plan-data" type="application/json">\s*(.*?)\s*</script>',
        text,
        re.S,
    )
    return json.loads(m.group(1))


def money_nums(s) -> list[float]:
    if s is None:
        return []
    out = []
    for x in re.findall(r"\$?\s*([\d,]+\.?\d*)", str(s)):
        x = x.replace(",", "").strip()
        if not x or x == ".":
            continue
        try:
            out.append(float(x))
        except ValueError:
            continue
    return out


def first_money(s) -> float | None:
    nums = money_nums(s)
    return nums[0] if nums else None


def pct_or_copay(s) -> tuple[str, float | None]:
    """Return ('pct'|'copay'|'lis'|'unknown', value)."""
    if s is None:
        return "unknown", None
    t = str(s).strip()
    if re.search(r"LIS|Extra Help|\$0-12\.65", t, re.I):
        return "lis", None
    if re.search(r"%", t) or (re.fullmatch(r"0\.\d+", t) and float(t) < 1):
        if "%" in t:
            m = re.search(r"([\d.]+)\s*%", t)
            return "pct", float(m.group(1)) / 100 if m else None
        return "pct", float(t)
    n = first_money(t)
    if n is not None:
        return "copay", n
    return "unknown", None


def normalize_giveback(v) -> float | None:
    if v is None:
        return None
    t = str(v).strip().lower()
    if t in {"", "no", "n/a", "na", "none", "not covered", "0", "0.0"}:
        return 0.0
    nums = money_nums(v)
    return nums[0] if nums else None


def parse_otc(text: str) -> dict | None:
    # Prefer explicit OTC credit/benefit lines (skip $0 noise / non-OTC hits)
    patterns = [
        r"(?:OTC(?:\s+credit)?|Over[- ]the[- ]Counter[^\n]{0,60}?)\s*(?:benefit)?[:\s]*\$\s*([\d,]+)\s*(?:credit\s+)?(?:every|/|per|x)?\s*(month|monthly|quarter|quarterly)",
        r"\$\s*([\d,]+)\s*(?:credit\s+)?(?:every|/|per|x)\s*(month|monthly|quarter|quarterly)[^\n]{0,60}(?:OTC|over[- ]the[- ]counter)",
        r"\$\s*([\d,]+)\s+monthly[^\n]{0,40}(?:OTC|over[- ]the[- ]counter)",
        r"(?:OTC|Over[- ]the[- ]Counter)[^\n]{0,80}?\$\s*([\d,]+)\s*(?:monthly|every\s+month|/ ?mo)",
        r"(?:OTC|Over[- ]the[- ]Counter)[^\n]{0,80}?\$\s*([\d,]+)\s*(?:every\s+)?quarter",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            amt = float(m.group(1).replace(",", ""))
            if amt <= 0:
                continue
            g2 = m.group(2) if m.lastindex and m.lastindex >= 2 else None
            blob = m.group(0).lower()
            if g2:
                period = "month" if str(g2).lower().startswith("month") else "quarter"
            else:
                period = "quarter" if "quarter" in blob else "month"
            return {"amount": amt, "period": period}
    if re.search(r"OTC[^\n]{0,40}(not covered|N/A|not available)", text, re.I):
        return {"amount": None, "period": None, "notCovered": True}
    return None


def parse_sob_fields(text: str) -> dict:
    """Best-effort field scrape across carrier SoB layouts."""
    t = text or ""
    out: dict = {}

    # Premium — monthly plan premium; capture ranges like $0 – $4.80
    m = re.search(
        r"(?:Monthly\s+Plan\s+Premium|Medical\s+premium|Monthly\s+plan\s+premium)\s*[:\n]?\s*(?:You\s+pay\s+)?\$\s*([\d,.]+)(?:\s*[-–/]\s*\$?\s*([\d,.]+))?",
        t,
        re.I,
    )
    if not m:
        m = re.search(
            r"Monthly\s+Plan\s+\$\s*([\d,.]+)(?:\s*[-–/]\s*\$?\s*([\d,.]+))?",
            t,
            re.I,
        )
    if m:
        a = float(m.group(1).replace(",", ""))
        b = float(m.group(2).replace(",", "")) if m.group(2) else None
        out["premium"] = {"low": a, "high": b or a, "raw": m.group(0)[:120]}

    # Part B giveback / reduction — require "$" so footnote markers aren't captured
    m = re.search(
        r"(?:Part\s*B\s*(?:Premium\s*)?(?:Reduction|give\s*back|giveback)|Part\s*B\s+premium\s+reduction).{0,120}?(?:[Uu]p\s+to\s+)?\$\s*([\d,.]+)",
        t,
        re.I | re.S,
    )
    if m:
        out["partBGiveback"] = float(m.group(1).replace(",", ""))
    elif re.search(
        r"Part\s*B\s*(?:Premium\s*)?(?:Reduction|give\s*back).{0,40}(not\s+available|N/?A|none|not\s+covered)",
        t,
        re.I | re.S,
    ):
        out["partBGiveback"] = 0.0

    # MOOP
    m = re.search(
        r"(?:Maximum\s+out[- ]of[- ]pocket|out[- ]of[- ]pocket\s+amount|yearly\s+limit)[^\n$]{0,120}?\$\s*([\d,]+)",
        t,
        re.I,
    )
    if not m:
        m = re.search(r"\$\s*([\d,]+)\s+for\s+services\s+you\s+receive", t, re.I)
    if m:
        out["moop"] = float(m.group(1).replace(",", ""))

    # PCP
    m = re.search(
        r"(?:Primary\s+care(?:\s+physician|\s+provider)?(?:\s+visit)?|PCP)[^\n]{0,50}?\$\s*([\d,]+)\s*(?:copay)?",
        t,
        re.I,
    )
    if m:
        out["pcpCopay"] = float(m.group(1).replace(",", ""))

    # Specialist
    m = re.search(
        r"Specialist(?:\s+visit)?[^\n]{0,50}?\$\s*([\d,]+)\s*(?:copay)?",
        t,
        re.I,
    )
    if m:
        out["specialistCopay"] = float(m.group(1).replace(",", ""))

    # ER — prefer the line immediately after "Emergency Care" (skip Worldwide)
    m = re.search(
        r"Emergency Care\s*\n\s*\$\s*([\d,.]+)\s*(?:\.00)?\s*copay",
        t,
        re.I,
    )
    if not m:
        m = re.search(
            r"Emergency(?:\s+care|\s+room)?(?![^\n]{0,40}[Ww]orldwide)[^\n]{0,40}?\$\s*([\d,]+)\s*(?:copay)?",
            t,
            re.I,
        )
    if m:
        out["erCopay"] = float(m.group(1).replace(",", ""))

    # Drug tiers (prefer first one-month retail block)
    for tier in range(1, 6):
        # "$0 copay" or "25% coinsurance" after Tier N
        m = re.search(
            rf"Tier\s*{tier}\s*[:\-]?\s*[^\n]{{0,60}}?(?:\$\s*([\d,.]+)\s*copay|([\d.]+)\s*%\s*coinsurance)",
            t,
            re.I,
        )
        if m:
            if m.group(1):
                out[f"tier{tier}"] = {"type": "copay", "value": float(m.group(1).replace(",", ""))}
            else:
                out[f"tier{tier}"] = {"type": "pct", "value": float(m.group(2)) / 100}

    otc = parse_otc(t)
    if otc:
        out["otc"] = otc

    return out


def max_premium_range(v) -> tuple[float | None, float | None]:
    if v is None:
        return None, None
    t = str(v)
    nums = money_nums(t)
    if not nums:
        return None, None
    if len(nums) == 1:
        return nums[0], nums[0]
    return min(nums), max(nums)


def max_otc(v) -> dict | None:
    if v is None:
        return None
    t = str(v)
    if re.search(r"not covered|n/?a", t, re.I):
        return {"amount": None, "period": None, "notCovered": True}
    m = re.search(r"\$?\s*([\d,.]+)\s*(?:x|/|per)\s*(month|monthly|quarter|quarterly)", t, re.I)
    if m:
        period = "month" if m.group(2).lower().startswith("month") else "quarter"
        return {"amount": float(m.group(1).replace(",", "")), "period": period}
    return None


def compare_plan(plan: dict, sob: dict) -> list[dict]:
    diffs = []
    pid = plan.get("planId") or plan.get("id")
    county = plan.get("county")

    def add(field, max_v, sob_v, severity, note=""):
        diffs.append(
            {
                "planId": pid,
                "county": county,
                "carrier": plan.get("carrier"),
                "planName": plan.get("planName"),
                "field": field,
                "maxValue": max_v,
                "sobValue": sob_v,
                "severity": severity,
                "note": note,
                "sobUrl": plan.get("sobUrl"),
            }
        )

    # Premium (skip LIS / Extra Help narrative cells — SoB often shows $0 member-pay line only)
    if "premium" in sob:
        prem_raw = str(plan.get("premium") or "")
        if re.search(r"LIS|Extra Help|Medicaid|Full:|Partial:|/\s*\$", prem_raw, re.I):
            pass
        else:
            lo, hi = max_premium_range(plan.get("premium"))
            s_lo, s_hi = sob["premium"]["low"], sob["premium"]["high"]
            if lo is not None and (abs(lo - s_lo) > 0.05 or abs((hi or lo) - s_hi) > 0.05):
                # DSNP-style $0 vs $0–$4.80: flag as med when Max is flat $0 and SoB has a small range
                sev = "high"
                if lo == 0 and s_lo == 0 and 0 < s_hi <= 5:
                    sev = "med"
                elif abs(lo - s_lo) < 1 and abs((hi or lo) - s_hi) < 1:
                    sev = "med"
                add(
                    "premium",
                    plan.get("premium"),
                    f"${s_lo:g}" + (f"–${s_hi:g}" if s_hi != s_lo else ""),
                    sev,
                )

    # Giveback
    if "partBGiveback" in sob:
        mv = normalize_giveback(plan.get("partBGiveback"))
        sv = sob["partBGiveback"]
        if mv is not None and abs(mv - sv) > 0.5:
            add("partBGiveback", plan.get("partBGiveback"), sv, "high")

    # MOOP
    if "moop" in sob:
        mv = first_money(plan.get("moop"))
        if mv is not None and abs(mv - sob["moop"]) > 1:
            add("moop", plan.get("moop"), sob["moop"], "high")

    for field in ("pcpCopay", "specialistCopay", "erCopay"):
        if field in sob:
            kind, mv = pct_or_copay(plan.get(field))
            if kind == "copay" and mv is not None and abs(mv - sob[field]) > 0.5:
                add(field, plan.get(field), sob[field], "high" if field != "pcpCopay" else "med")

    for tier in range(1, 6):
        key = f"tier{tier}"
        if key not in sob:
            continue
        kind, mv = pct_or_copay(plan.get(key))
        sv = sob[key]
        if kind == "lis":
            continue  # DSNP LIS framing — don't false-flag vs coinsurance lines
        if kind == "copay" and sv["type"] == "copay" and mv is not None and abs(mv - sv["value"]) > 0.5:
            add(key, plan.get(key), sv["value"], "high")
        if kind == "pct" and sv["type"] == "pct" and mv is not None and abs(mv - sv["value"]) > 0.015:
            add(key, plan.get(key), f"{sv['value']*100:g}%", "high")
        if kind == "copay" and sv["type"] == "pct":
            add(key, plan.get(key), f"{sv['value']*100:g}%", "med", "Max copay vs SoB %")
        if kind == "pct" and sv["type"] == "copay":
            add(key, plan.get(key), sv["value"], "med", "Max % vs SoB copay")

    if "otc" in sob:
        mo = max_otc(plan.get("otc"))
        so = sob["otc"]
        if mo and so.get("notCovered") and not mo.get("notCovered"):
            add("otc", plan.get("otc"), "Not covered", "med")
        elif mo and so.get("amount") is not None and mo.get("amount") is not None:
            if abs(mo["amount"] - so["amount"]) > 0.5 or mo.get("period") != so.get("period"):
                add(
                    "otc",
                    plan.get("otc"),
                    f"${so['amount']:g} x {so['period']}",
                    "high",
                )

    return diffs


def main() -> None:
    plans = load_plans()
    extracted = json.loads(EXTRACTED.read_text(encoding="utf-8"))
    fields_out = {}
    all_diffs = []
    extract_stats = {"ok": 0, "fail": 0, "parsed": 0}

    plan_by_key = {(p.get("planId") or p.get("id"), p.get("county")): p for p in plans}

    for key, rec in extracted.items():
        if "|" in key:
            pid, county = key.split("|", 1)
        else:
            pid, county = rec.get("planId"), rec.get("county")
        plan = plan_by_key.get((pid, county))
        if not plan:
            # fuzzy
            cands = [p for p in plans if (p.get("planId") or p.get("id")) == pid]
            plan = cands[0] if len(cands) == 1 else None
        if not rec.get("text"):
            extract_stats["fail"] += 1
            continue
        extract_stats["ok"] += 1
        sob = parse_sob_fields(rec["text"])
        fields_out[key] = sob
        if sob:
            extract_stats["parsed"] += 1
        if plan:
            all_diffs.extend(compare_plan(plan, sob))

    OUT_JSON.write_text(json.dumps(fields_out, indent=2), encoding="utf-8")

    # CSV
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "severity",
                "planId",
                "county",
                "carrier",
                "planName",
                "field",
                "maxValue",
                "sobValue",
                "note",
                "sobUrl",
            ],
        )
        w.writeheader()
        for d in sorted(all_diffs, key=lambda x: (0 if x["severity"] == "high" else 1, x["planId"], x["field"])):
            w.writerow(d)

    # XLSX workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corrections"
    headers = [
        "Severity",
        "Plan ID",
        "County",
        "Carrier",
        "Plan Name",
        "Field",
        "Max Value",
        "SoB Value",
        "Note",
        "SoB URL",
        "Action",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    fills = {
        "high": PatternFill("solid", fgColor="F8D7DA"),
        "med": PatternFill("solid", fgColor="FFF3CD"),
    }
    for d in sorted(all_diffs, key=lambda x: (0 if x["severity"] == "high" else 1, x["planId"], x["field"])):
        row = [
            d["severity"],
            d["planId"],
            d["county"],
            d["carrier"],
            d["planName"],
            d["field"],
            d["maxValue"],
            d["sobValue"],
            d["note"],
            d["sobUrl"],
            "review",
        ]
        ws.append(row)
        fill = fills.get(d["severity"])
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    ws2 = wb.create_sheet("Extract Stats")
    ws2.append(["Metric", "Value"])
    ws2.append(["Extracts with text", extract_stats["ok"]])
    ws2.append(["Extract failures", extract_stats["fail"]])
    ws2.append(["Parsed ≥1 field", extract_stats["parsed"]])
    ws2.append(["Diff rows", len(all_diffs)])
    ws2.append(["High severity", sum(1 for d in all_diffs if d["severity"] == "high")])
    ws2.append(["Med severity", sum(1 for d in all_diffs if d["severity"] == "med")])

    # Suggested auto-fixes (clear high-confidence)
    ws3 = wb.create_sheet("Suggested Auto-Fixes")
    ws3.append(headers)
    auto = []
    for d in all_diffs:
        if d["severity"] != "high":
            continue
        # Only suggest when SoB value is a simple number/range we can write
        if d["field"] in {"moop", "pcpCopay", "specialistCopay", "erCopay", "partBGiveback", "premium", "otc"} or d[
            "field"
        ].startswith("tier"):
            auto.append(d)
            row = [
                d["severity"],
                d["planId"],
                d["county"],
                d["carrier"],
                d["planName"],
                d["field"],
                d["maxValue"],
                d["sobValue"],
                d["note"],
                d["sobUrl"],
                "candidate",
            ]
            ws3.append(row)

    wb.save(OUT_XLSX)
    print(
        f"extract_ok={extract_stats['ok']} fail={extract_stats['fail']} "
        f"parsed={extract_stats['parsed']} diffs={len(all_diffs)} "
        f"high={sum(1 for d in all_diffs if d['severity']=='high')} "
        f"-> {OUT_XLSX}"
    )


if __name__ == "__main__":
    main()
