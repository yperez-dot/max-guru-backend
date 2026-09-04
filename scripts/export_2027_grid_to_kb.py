#!/usr/bin/env python3
"""Export confirmed (green) 2027 THEI grid cells into max-knowledge markdown.

Yellow cells are leftover 2026 numbers — they are never written as 2027 facts.
Re-run after each sheet refresh. Does not touch live #plan-data (stays 2026).
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

SHEET_ID = "1BYhBfOzdeJOMEVXIKJkHrZzEohrOBR-N"
XLSX_PATH = Path("/tmp/thei-2027-grid.xlsx")
KB_DIR = Path("/workspace/max-knowledge/carriers")
OVERVIEW_PATH = KB_DIR / "plan-grid-overview-2027.md"
WATCH_PATH = Path("/workspace/artifacts/reports/2027-grid-watch-state.json")

SHEETS = [
    ("DADE- HMO", "Miami-Dade", "HMO"),
    ("BWD- HMO", "Broward", "HMO"),
    ("DADE-CSNP", "Miami-Dade", "C-SNP"),
    ("BWD-CSNP", "Broward", "C-SNP"),
    ("Dade- DSNP", "Miami-Dade", "D-SNP"),
    ("BWD-DSNP", "Broward", "D-SNP"),
    ("DADE- Giveback", "Miami-Dade", "Giveback"),
    ("BWD-Giveback", "Broward", "Giveback"),
    ("DADE-PPO", "Miami-Dade", "PPO"),
    ("BWD-PPO", "Broward", "PPO"),
]

CARRIER_FILES = {
    "Humana": "humana-plans-florida-2027.md",
    "Devoted": "devoted-plans-florida-2027.md",
    "UHC": "uhc-plans-florida-2027.md",
    "CarePlus": "careplus-plans-florida-2027.md",
    "Aetna": "aetna-plans-florida-2027.md",
}

CARRIER_ALIASES = [
    ("UHC", "UHC"),
    ("United", "UHC"),
    ("Preferred", "UHC"),
    ("AARP", "UHC"),
    ("MedicareMax", "UHC"),
    ("CarePlus", "CarePlus"),
    ("CareOne", "CarePlus"),
    ("CareNeeds", "CarePlus"),
    ("CareFree", "CarePlus"),
    ("CareBreeze", "CarePlus"),
    ("CareComplete", "CarePlus"),
    ("CareAccess", "CarePlus"),
    ("Devoted", "Devoted"),
    ("Aetna", "Aetna"),
    ("Humana", "Humana"),
    ("Simply", "Simply"),
    ("HealthSun", "HealthSun"),
    ("Wellcare", "Wellcare"),
    ("WellCare", "Wellcare"),
    ("Doctors", "Doctors"),
    ("Doctor", "Doctors"),
    ("Florida Blue", "Florida Blue"),
    ("FL Blue", "Florida Blue"),
    ("HealthSpring", "HealthSpring"),
    ("Cigna", "HealthSpring"),
    ("Solis", "Solis"),
    ("Gold Kidney", "Gold Kidney"),
]


def fill_rgb(cell) -> str | None:
    fill = cell.fill
    if not fill or fill.fill_type is None:
        return None
    fg = fill.fgColor
    if fg is None or fg.type != "rgb" or not fg.rgb:
        return None
    return str(fg.rgb).upper()


def is_green(cell) -> bool:
    rgb = fill_rgb(cell)
    return bool(rgb and (rgb.endswith("E8F5E9") or rgb.endswith("C8E6C9")))


def is_yellow(cell) -> bool:
    rgb = fill_rgb(cell)
    return bool(
        rgb
        and (
            rgb.endswith("FFF2CC")
            or rgb.endswith("FFFF00")
            or rgb.endswith("FFEB3B")
        )
    )


def extract_plan_id(header: object) -> str | None:
    if not header:
        return None
    s = str(header).replace("‑", "-").replace("–", "-")
    m = re.search(r"\b([HR]\d{3,4})\s*\|\s*(\d{2,4})\b", s, re.I)
    if m:
        return f"{m.group(1).upper()}-{m.group(2)}"
    m = re.search(
        r"\b([HR]\d{3,4})\s*-\s*(\d{2,4}[A-Z]?)(?:\s*(?:FL-?)(\d{2,4})|\s*-\s*(\d{1,4})|\s*/\s*-?\s*(\d{2,4}))?",
        s,
        re.I,
    )
    if not m:
        m = re.search(
            r"\b([HR]\d{3})\s*-\s*(\d{2,4}[A-Z]?)(?:\s*-\s*(\d{1,4}))?",
            s,
            re.I,
        )
        if not m:
            return None
        base = f"{m.group(1).upper()}-{m.group(2)}"
        if m.lastindex and m.lastindex >= 3 and m.group(3):
            return f"{base}-{m.group(3)}"
        return base
    base = f"{m.group(1).upper()}-{m.group(2)}"
    if m.group(3):
        return f"{base}-FL-{m.group(3)}"
    if m.group(4):
        return f"{base}-{m.group(4)}"
    if m.group(5):
        return f"{base}/{m.group(5)}"
    return base


def carrier_of(header: str) -> str:
    for key, nice in CARRIER_ALIASES:
        if re.search(rf"\b{re.escape(key)}\b", header, re.I) or header.startswith(key):
            return nice
    return header.split()[0] if header.split() else "Unknown"


def clean_plan_name(header: str, pid: str | None) -> str:
    s = header.replace("‑", "-").replace("–", "-")
    s = re.sub(r"\s+", " ", s.replace("\n", " ")).strip()
    s = re.sub(r"\*{1,2}[^*]*\*{1,2}", " ", s)
    s = re.sub(r"\bNON[- ]COMMI?SSIONABLE\b", "", s, flags=re.I)
    s = re.sub(r"\bNEW\s*2027\b", "", s, flags=re.I)
    s = re.sub(r"\bCLOSED(?:\s+TO)?\s+NEW\s+ENROLL(?:MENT)?(?:\s+2027)?\b", "", s, flags=re.I)
    if pid:
        s = re.sub(re.escape(pid), "", s, flags=re.I)
        compact = pid.replace("-", r"\s*[|\-]?\s*")
        s = re.sub(compact, "", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip(" -|")
    return s or (pid or "Unknown")


def fmt_val(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    s = str(val).replace("\n", " ").strip()
    s = re.sub(r"\s+", " ", s)
    if s in {'"', "“", "”"}:
        return ""
    return s


def sob_target(cell) -> str | None:
    if cell.hyperlink and cell.hyperlink.target:
        return str(cell.hyperlink.target).strip()
    if isinstance(cell.value, str) and cell.value.strip().startswith("http"):
        return cell.value.strip()
    return None


def header_flags(header: str) -> list[str]:
    flags = []
    h = header.upper()
    if "NEW 2027" in h or "**NEW" in h:
        flags.append("NEW 2027")
    if "CLOSED" in h:
        flags.append("CLOSED TO NEW ENROLL 2027")
    if "NON-COMM" in h or "NON COMM" in h:
        flags.append("NON-COMMISSIONABLE (new sales)")
    return flags


def parse_grid(xlsx: Path) -> tuple[list[dict], dict]:
    wb = openpyxl.load_workbook(xlsx, data_only=False)
    notes_ws = wb["2027 NOTES"] if "2027 NOTES" in wb.sheetnames else None
    notes = []
    if notes_ws:
        for r in range(1, 30):
            v = notes_ws.cell(r, 1).value
            if v:
                notes.append(str(v).strip())

    plans: list[dict] = []
    for sheet, county, ptype in SHEETS:
        ws = wb[sheet]
        labels: list[tuple[int, str]] = []
        sob_row = None
        for r in range(1, (ws.max_row or 0) + 1):
            lab = ws.cell(r, 1).value
            if not lab or not str(lab).strip():
                continue
            lab_s = re.sub(r"\s+", " ", str(lab).replace("\n", " ")).strip()
            if lab_s.lower().startswith("summary of"):
                sob_row = r
                continue
            if lab_s.lower() in {"note", "star ratings", "star rating"}:
                continue
            labels.append((r, lab_s))

        max_col = ws.max_column or 0
        for c in range(2, max_col + 1):
            header_cell = ws.cell(1, c)
            header = header_cell.value
            if header is None or not str(header).strip():
                continue
            header_s = str(header)
            pid = extract_plan_id(header_s)
            if not pid:
                continue
            carrier = carrier_of(header_s)
            name = clean_plan_name(header_s, pid)

            fields_green: list[tuple[str, str]] = []
            fields_yellow = 0
            fields_other = 0
            for r, lab in labels:
                cell = ws.cell(r, c)
                val = fmt_val(cell.value)
                if not val:
                    # PPO IN/OUT: if this is an IN column, also check the next OUT col
                    continue
                if is_green(cell):
                    fields_green.append((lab, val))
                elif is_yellow(cell):
                    fields_yellow += 1
                else:
                    fields_other += 1

            # Pair PPO OUT column (header blank, row 2 often "OUT")
            out_fields: list[tuple[str, str]] = []
            if ptype == "PPO" and c + 1 <= max_col and not ws.cell(1, c + 1).value:
                for r, lab in labels:
                    cell = ws.cell(r, c + 1)
                    val = fmt_val(cell.value)
                    if val and is_green(cell):
                        out_fields.append((lab, val))
                    elif val and is_yellow(cell):
                        fields_yellow += 1

            sob_cell = ws.cell(sob_row, c) if sob_row else None
            sob_url = sob_target(sob_cell) if sob_cell else None
            sob_caption = fmt_val(sob_cell.value) if sob_cell else ""
            sob_confirmed = bool(sob_cell and (is_green(sob_cell) or sob_caption.lower().startswith("2027")))

            if not fields_green and not sob_confirmed:
                continue

            plans.append(
                {
                    "id": pid,
                    "carrier": carrier,
                    "planName": name,
                    "county": county,
                    "type": ptype,
                    "sheet": sheet,
                    "header": re.sub(r"\s+", " ", header_s.replace("\n", " ")).strip(),
                    "flags": header_flags(header_s),
                    "fields": fields_green,
                    "outFields": out_fields,
                    "yellowLeft": fields_yellow,
                    "sobUrl": sob_url,
                    "sobCaption": sob_caption,
                    "sobConfirmed": sob_confirmed,
                }
            )

    meta = {
        "notes": notes,
        "sheetModified": str(wb.properties.modified) if wb.properties.modified else None,
    }
    return plans, meta


def md_escape(s: str) -> str:
    return s.replace("|", "\\|")


def render_plan(p: dict) -> str:
    flags = f" — {' · '.join(p['flags'])}" if p["flags"] else ""
    lines = [
        f"## {p['planName']} ({p['id']}) — {p['county']} {p['type']}{flags}",
        f"**County:** {p['county']}",
        f"**Type:** {p['type']}",
        f"**Plan year:** 2027",
        f"**CMS ID:** {p['id']}",
    ]
    if p["sobUrl"] and p["sobConfirmed"]:
        lines.append(f"**SoB:** [SoB]({p['sobUrl']})")
        if p["sobCaption"] and not p["sobCaption"].lower().startswith("summary of"):
            lines.append(f"**SoB note:** {p['sobCaption']}")
    elif p["sobCaption"] and p["sobConfirmed"]:
        lines.append(f"**SoB note:** {p['sobCaption']} (no URL on the grid cell)")
    if p["yellowLeft"]:
        lines.append(
            f"**Still yellow on the working grid:** {p['yellowLeft']} field(s) — not cited below."
        )
    # highlight a few common fields at the top when green
    top = {k.lower(): v for k, v in p["fields"]}
    for label, key in (
        ("premium", "Premium"),
        ("part b giveback", "Part B Giveback"),
        ("part b give back", "Part B Giveback"),
        ("part b rebate", "Part B Giveback"),
        ("max out of pocket", "MOOP"),
    ):
        if label in top:
            lines.append(f"**{key}:** {top[label]}")
    lines.append("")
    lines.append("| Benefit | 2027 (confirmed) |")
    lines.append("|---------|------------------|")
    for lab, val in p["fields"]:
        lines.append(f"| {md_escape(lab)} | {md_escape(val)} |")
    if p["outFields"]:
        lines.append("")
        lines.append("Out-of-network (confirmed):")
        lines.append("")
        lines.append("| Benefit | 2027 OON |")
        lines.append("|---------|----------|")
        for lab, val in p["outFields"]:
            lines.append(f"| {md_escape(lab)} | {md_escape(val)} |")
    lines.append("")
    return "\n".join(lines)


def render_carrier(carrier: str, plans: list[dict], pulled: str) -> str:
    plans = sorted(plans, key=lambda p: (p["county"], p["type"], p["id"]))
    ids = ", ".join(sorted({p["id"] for p in plans}))
    counties = " / ".join(sorted({p["county"] for p in plans}))
    body = [
        f"# {carrier} — Florida 2027 plans",
        f"Source: THEI 2027 Plan Benefit Grid (working copy) — confirmed green cells only.",
        f"Workbook: https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit",
        f"Pulled: {pulled}",
        f"Counties: {counties}",
        f"CMS IDs on file: {ids}",
        "",
        "Yellow leftover 2026 cells are **not** in this file. If a benefit is missing here, Max does not have a confirmed 2027 figure yet.",
        "",
    ]
    for p in plans:
        body.append(render_plan(p))
        body.append("---")
        body.append("")
    return "\n".join(body).rstrip() + "\n"


def render_overview(plans: list[dict], meta: dict, pulled: str, stats: dict) -> str:
    by_carrier = Counter(p["carrier"] for p in plans)
    confirmed = [p for p in plans if p["fields"] or p["sobConfirmed"]]
    new_plans = [p for p in plans if "NEW 2027" in p["flags"]]
    closed = [p for p in plans if any("CLOSED" in f for f in p["flags"])]
    noncomm = [p for p in plans if any("NON-COMM" in f for f in p["flags"])]

    waiting = [
        "Doctors",
        "Florida Blue",
        "HealthSpring / Cigna",
        "HealthSun",
        "Simply",
        "Solis",
        "Wellcare",
        "Gold Kidney",
    ]

    lines = [
        "# 2027 THEI plan grid — what Max can cite",
        f"Source: THEI 2027 Plan Benefit Grid working copy ([Google Sheet](https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit))",
        f"Pulled: {pulled}",
        f"Sheet stamp: {stats.get('green_cells')} green / {stats.get('yellow_cells')} yellow benefit cells across plan tabs.",
        "",
        "Color key on the sheet: **light green** = 2027 number from an official SoB, highlight, or sneak-peek slide (official SoB wins). **Yellow** = still the 2026 number. Max only cites green.",
        "",
        "Live `#plan-data` stays the **2026** grid. These files are how chat answers 2027.",
        "",
        "## Confirmed 2027 plan dollars (green)",
        "",
        "| Carrier | Plans with confirmed 2027 cells | KB doc |",
        "|---------|----------------------------------|--------|",
    ]
    for carrier, fname in CARRIER_FILES.items():
        n = by_carrier.get(carrier, 0)
        key = fname.replace(".md", "")
        lines.append(f"| {carrier} | {n} | `carriers/{key}` |")
    lines.append("")
    lines.append("## Still waiting on the official October 1 SoB")
    lines.append("")
    lines.append(
        "These carriers are on the 2027 workbook but every benefit cell is still yellow. "
        "Do **not** quote their 2026 leftover numbers as 2027. Say Max does not have that 2027 figure yet."
    )
    lines.append("")
    for w in waiting:
        lines.append(f"- {w}")
    lines.append("")
    if new_plans:
        lines.append("## New 2027 plans on the grid")
        lines.append("")
        seen = set()
        for p in sorted(new_plans, key=lambda x: (x["id"], x["county"])):
            key = (p["id"], p["county"])
            if key in seen:
                continue
            seen.add(key)
            lines.append(f"- {p['planName']} (`{p['id']}`) — {p['county']} {p['type']}")
        lines.append("")
    if closed:
        lines.append("## Closed to new enroll 2027")
        lines.append("")
        seen = set()
        for p in sorted(closed, key=lambda x: (x["id"], x["county"])):
            key = (p["id"], p["county"])
            if key in seen:
                continue
            seen.add(key)
            lines.append(f"- {p['planName']} (`{p['id']}`) — {p['county']}")
        lines.append("")
    if noncomm:
        lines.append("## Marked non-commissionable on the 2027 grid (new sales)")
        lines.append("")
        seen = set()
        for p in noncomm:
            if p["id"] in seen:
                continue
            seen.add(p["id"])
            lines.append(f"- {p['planName']} (`{p['id']}`)")
        lines.append("")
    lines.append("## Hospital / network notes already confirmed for 2027")
    lines.append("")
    lines.append(
        "- **UHealth / University of Miami** and **Bascom Palmer** are **out of MedicareMax (Preferred Care Network)** as of **1/1/2027**. In-network through 12/31/2026. University Hospital on the Hospitals tab is HCA Davie — not UM."
    )
    lines.append(
        "- Other hospital Yes/— marks stay 2026 until a public 2027 directory lands (due Oct 1, 2026)."
    )
    lines.append("")
    lines.append("## Workbook notes (from the 2027 NOTES tab)")
    lines.append("")
    for n in meta.get("notes") or []:
        lines.append(f"- {n}")
    lines.append("")
    lines.append("## How to refresh")
    lines.append("")
    lines.append("```bash")
    lines.append(
        f"curl -sL -o /tmp/thei-2027-grid.xlsx "
        f"'https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx'"
    )
    lines.append("python3 scripts/export_2027_grid_to_kb.py")
    lines.append("```")
    lines.append("")
    return "\n".join(lines)


def count_fills(xlsx: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx, data_only=False)
    green = yellow = 0
    for sheet, _, _ in SHEETS:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row or 0, 80), min_col=2):
            for cell in row:
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                if is_green(cell):
                    green += 1
                elif is_yellow(cell):
                    yellow += 1
    return {"green_cells": green, "yellow_cells": yellow}


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"missing {XLSX_PATH}", file=sys.stderr)
        return 1
    pulled = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    digest = hashlib.sha256(XLSX_PATH.read_bytes()).hexdigest()
    plans, meta = parse_grid(XLSX_PATH)
    stats = count_fills(XLSX_PATH)

    by_carrier: dict[str, list[dict]] = defaultdict(list)
    for p in plans:
        if p["carrier"] in CARRIER_FILES:
            by_carrier[p["carrier"]].append(p)

    written = []
    for carrier, fname in CARRIER_FILES.items():
        path = KB_DIR / fname
        text = render_carrier(carrier, by_carrier.get(carrier, []), pulled)
        path.write_text(text, encoding="utf-8")
        written.append(str(path))
        print(f"wrote {path} ({len(by_carrier.get(carrier, []))} plans)")

    OVERVIEW_PATH.write_text(render_overview(plans, meta, pulled, stats), encoding="utf-8")
    written.append(str(OVERVIEW_PATH))
    print(f"wrote {OVERVIEW_PATH}")

    WATCH_PATH.parent.mkdir(parents=True, exist_ok=True)
    state = {
        "sheetId": SHEET_ID,
        "pulled": pulled,
        "sha256": digest,
        "bytes": XLSX_PATH.stat().st_size,
        "confirmedPlans": len(plans),
        "byCarrier": {k: len(v) for k, v in by_carrier.items()},
        "greenCells": stats["green_cells"],
        "yellowCells": stats["yellow_cells"],
        "planIds": sorted({p["id"] for p in plans}),
        "new2027": [p["id"] for p in plans if "NEW 2027" in p["flags"]],
        "closed2027": [p["id"] for p in plans if any("CLOSED" in f for f in p["flags"])],
    }
    WATCH_PATH.write_text(json.dumps(state, indent=2) + "\n", encoding="utf-8")
    print(f"watch {WATCH_PATH} sha256={digest[:12]} plans={len(plans)} green={stats['green_cells']} yellow={stats['yellow_cells']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
